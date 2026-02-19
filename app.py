from __future__ import annotations

import json
from pathlib import Path
from datetime import datetime

from PySide6.QtCore import *
from PySide6.QtGui import *
from PySide6.QtWidgets import *


from afip_service import AfipService, CBTE_TIPO_FACTURA_C, CBTE_TIPO_FACTURA_B
from db import (
    init_db, insert_invoice, list_invoices, daily_summary, range_summary,
    get_invoice_with_items, get_all_settings, set_setting, get_setting,
    get_ticket_lines, save_ticket_lines, validate_license,
    normalize_taxpayer_type, is_cbte_allowed_for_taxpayer, allowed_cbte_types_for_taxpayer,
    default_cbte_for_taxpayer, taxpayer_type_lock_text, taxpayer_blocked_cbte_message
)
from ticket_format import build_ticket_text
from printer import TicketPrinter
from qr_afip import build_afip_qr_payload, build_afip_qr_url
from qr_debug import log_qr
from qr_render import make_qr_image

DOC_TIPO_SIN_DOC = 99
DOC_TIPO_DNI = 96
DOC_TIPO_CUIT = 80

def load_boot_settings() -> dict:
    """
    Archivo mínimo (solo para saber db_path).
    Todo lo demás vive en SQLite (Mandamiento #1).
    """
    p = Path("settings.json")
    if not p.exists():
        p.write_text(json.dumps({"db_path": "data/locutorio.sqlite"}, indent=2), encoding="utf-8")
    return json.loads(p.read_text(encoding="utf-8"))

def client_label(doc_tipo: int, doc_nro: str) -> str:
    if int(doc_tipo) == DOC_TIPO_SIN_DOC or str(doc_nro) in ("0", ""):
        return "A CONS FINAL"
    if int(doc_tipo) == DOC_TIPO_CUIT:
        return f"CUIT {doc_nro}"
    if int(doc_tipo) == DOC_TIPO_DNI:
        return f"DNI {doc_nro}"
    return f"DOC {doc_nro}"

def print_ticket_from_data(*, db_path: str, inv: dict, items: list[dict]) -> None:
    settings = get_all_settings(db_path)

    pv = int(inv["pv"])
    cbte_tipo = int(inv["cbte_tipo"])
    cbte_nro = int(inv["cbte_nro"])
    cae = str(inv["cae"])
    cae_vto = str(inv["cae_vto"])
    doc_tipo = int(inv["doc_tipo"])
    doc_nro = str(inv["doc_nro"])
    total = float(inv["imp_total"])
    modo = str(inv["modo"])

    now = datetime.now()

    ticket_text = build_ticket_text(
        db_path=db_path,
        cbte_tipo_label="FACTURA B" if cbte_tipo == CBTE_TIPO_FACTURA_B else "FACTURA C",
        pv=pv,
        cbte_nro=cbte_nro,
        fecha=now,
        items=[{"name": it["item_name"], "qty": it["qty"], "price": it["price"], "subtotal": it["subtotal"]} for it in items],
        total=total,
        cae=cae,
        cae_vto_yyyymmdd=cae_vto,
        cliente_label=client_label(doc_tipo, doc_nro),
    )

    # QR + debug
    is_cf = (int(doc_tipo) == DOC_TIPO_SIN_DOC) or (str(doc_nro) in ("0", ""))
    payload = build_afip_qr_payload(
        fecha_emision=now,
        cuit_emisor=int(settings.get("cuit_emisor") or 0),
        pto_vta=pv,
        tipo_cmp=int(cbte_tipo),
        nro_cmp=cbte_nro,
        importe=total,
        moneda="PES",
        ctz=1.0,
        tipo_cod_aut="E",
        cod_aut=cae,  # <-- CAE
        tipo_doc_rec=None if is_cf else int(doc_tipo),
        nro_doc_rec=None if is_cf else str(doc_nro),
    )
    qr_url = build_afip_qr_url(
        fecha_emision=now,
        cuit_emisor=int(settings.get("cuit_emisor") or 0),
        pto_vta=pv,
        tipo_cmp=int(cbte_tipo),
        nro_cmp=cbte_nro,
        importe=total,
        moneda="PES",
        ctz=1.0,
        tipo_cod_aut="E",
        cod_aut=cae,
        tipo_doc_rec=None if is_cf else int(doc_tipo),
        nro_doc_rec=None if is_cf else str(doc_nro),
    )
    log_qr(path="qr.log", url=qr_url, payload=payload)

    printer = TicketPrinter(settings.get("printer_name_contains") or None, mode=settings.get("print_mode", "escpos"))
    qr_img = make_qr_image(qr_url, box_size=4, border=2)
    printer.print_text_and_qr(ticket_text, qr_img)

class SetupWizard(QWidget):
    """
    'Runonce' simple: si setup_completed=0, mostramos esta pantalla.
    """
    def __init__(self, db_path: str):
        super().__init__()
        self.db_path = db_path
        self.setWindowTitle("Configuración inicial")
        lay = QVBoxLayout(self)

        form = QFormLayout()
        lay.addLayout(form)

        self.ed_app = QLineEdit(get_setting(db_path, "app_name", "LocutorioWEB"))
        self.ed_rs = QLineEdit(get_setting(db_path, "razon_social", "LocutorioWEB"))
        self.ed_cuit = QLineEdit(get_setting(db_path, "cuit_emisor", "0"))
        self.ed_pv = QLineEdit(get_setting(db_path, "punto_venta", "14"))

        self.cmb_modo = QComboBox()
        self.cmb_modo.addItem("Producción (real)", "PROD")
        self.cmb_modo.addItem("Homologación (pruebas)", "HOMO")
        cur_modo = (get_setting(db_path, "modo", "PROD") or "PROD").strip().upper()
        idx = 0 if cur_modo == "PROD" else 1
        self.cmb_modo.setCurrentIndex(idx)

        self.cmb_taxpayer = QComboBox()
        self.cmb_taxpayer.addItem("Monotributo (solo Factura C)", "MONO")
        self.cmb_taxpayer.addItem("Responsable Inscripto (solo Factura B)", "RI")
        cur_taxpayer = (get_setting(db_path, "taxpayer_type", "MONO") or "MONO").strip().upper()
        self.cmb_taxpayer.setCurrentIndex(0 if cur_taxpayer != "RI" else 1)

        self.ed_printer_contains = QLineEdit(get_setting(db_path, "printer_name_contains", ""))
        self.cmb_print_mode = QComboBox()
        self.cmb_print_mode.addItem("Térmica USB (XP58/Zebra) - ESC/POS", "escpos")
        self.cmb_print_mode.addItem("Impresora Windows (PDF/Láser) - GDI", "gdi")
        cur_pm = (get_setting(db_path, "print_mode", "escpos") or "escpos").strip().lower()
        self.cmb_print_mode.setCurrentIndex(0 if cur_pm == "escpos" else 1)

        self.ed_openssl = QLineEdit(get_setting(db_path, "openssl_path", ""))
        self.ed_crt = QLineEdit(get_setting(db_path, "cert_crt_path", ""))
        self.ed_key = QLineEdit(get_setting(db_path, "private_key_path", ""))
        self.ed_keypass = QLineEdit(get_setting(db_path, "private_key_password", ""))
        self.ed_keypass.setEchoMode(QLineEdit.Password)

        form.addRow("Nombre app:", self.ed_app)
        form.addRow("Razón social:", self.ed_rs)
        form.addRow("CUIT emisor:", self.ed_cuit)
        form.addRow("Punto de venta:", self.ed_pv)
        form.addRow("Modo AFIP:", self.cmb_modo)
        form.addRow("Régimen fiscal:", self.cmb_taxpayer)
        form.addRow("", QLabel("• Producción = emite comprobantes reales. Homologación = pruebas."))
        form.addRow("Impresora (opcional):", self.ed_printer_contains)
        form.addRow("", QLabel("• Dejá vacío para usar la impresora predeterminada de Windows."))
        form.addRow("Tipo de impresión:", self.cmb_print_mode)
        form.addRow("", QLabel("• ESC/POS: manda comandos directos a térmicas. GDI: sirve para PDF y cualquier driver Windows."))
        form.addRow("OpenSSL path:", self.ed_openssl)
        form.addRow("Cert .crt path:", self.ed_crt)
        form.addRow("Private key path:", self.ed_key)
        form.addRow("Key password:", self.ed_keypass)

        row = QHBoxLayout()
        lay.addLayout(row)
        self.btn_save = QPushButton("Guardar y continuar")
        self.btn_cancel = QPushButton("Cancelar")
        row.addStretch(1)
        row.addWidget(self.btn_cancel)
        row.addWidget(self.btn_save)

        self.btn_cancel.clicked.connect(self.close)
        self.btn_save.clicked.connect(self.on_save)

    def on_save(self):
        set_setting(self.db_path, "app_name", self.ed_app.text().strip())
        set_setting(self.db_path, "razon_social", self.ed_rs.text().strip())
        set_setting(self.db_path, "cuit_emisor", self.ed_cuit.text().strip())
        set_setting(self.db_path, "punto_venta", self.ed_pv.text().strip())
        set_setting(self.db_path, "modo", str(self.cmb_modo.currentData()).strip())
        set_setting(self.db_path, "taxpayer_type", str(self.cmb_taxpayer.currentData()).strip())
        set_setting(self.db_path, "printer_name_contains", self.ed_printer_contains.text().strip())
        set_setting(self.db_path, "print_mode", str(self.cmb_print_mode.currentData()).strip())
        set_setting(self.db_path, "openssl_path", self.ed_openssl.text().strip())
        set_setting(self.db_path, "cert_crt_path", self.ed_crt.text().strip())
        set_setting(self.db_path, "private_key_path", self.ed_key.text().strip())
        set_setting(self.db_path, "private_key_password", self.ed_keypass.text())
        set_setting(self.db_path, "setup_completed", "1")
        QMessageBox.information(self, "OK", "Configuración guardada.")
        self.close()

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()

        boot = load_boot_settings()
        self.db_path = boot["db_path"]

        init_db(self.db_path)

        # licencia (si enabled=0 no molesta)
        validate_license(self.db_path)

        # runonce
        if get_setting(self.db_path, "setup_completed", "0") != "1":
            wiz = SetupWizard(self.db_path)
            wiz.destroyed.connect(lambda *_: self._refresh_after_setup())
            wiz.show()
            # seguimos igual; el usuario puede configurar luego en la pestaña

        self.settings = get_all_settings(self.db_path)
        self.setWindowTitle(f"{self.settings.get('app_name','Factura')}")

        self.afip = AfipService(
            modo=self.settings.get("modo", "PROD"),
            pv=int(self.settings.get("punto_venta", "14") or 14),
            cuit=int(self.settings.get("cuit_emisor", "0") or 0),
            cert_crt_path=self.settings.get("cert_crt_path", ""),
            private_key_path=self.settings.get("private_key_path", ""),
            private_key_password=(self.settings.get("private_key_password") or None),
            openssl_path=(self.settings.get("openssl_path") or None),
        )

        layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        self._build_tab_venta()
        self._build_tab_comprobantes()
        self._build_tab_reporte()
        self._build_tab_config()

        # Shortcut F4 => facturar+imprimir
        QShortcut(QKeySequence("F4"), self, activated=self.on_facturar)

    def _is_cbte_allowed(self, cbte_tipo: int) -> bool:
        taxpayer_type = get_setting(self.db_path, "taxpayer_type", "MONO")
        return is_cbte_allowed_for_taxpayer(taxpayer_type=taxpayer_type, cbte_tipo=cbte_tipo)

    def _apply_cbte_lock(self) -> None:
        taxpayer_type = normalize_taxpayer_type(get_setting(self.db_path, "taxpayer_type", "MONO"))
        allowed = set(allowed_cbte_types_for_taxpayer(taxpayer_type))

        self.cbte_combo.blockSignals(True)
        try:
            self.cbte_combo.clear()
            if CBTE_TIPO_FACTURA_C in allowed:
                self.cbte_combo.addItem("Factura C (Monotributo / no discrimina IVA)", CBTE_TIPO_FACTURA_C)
            if CBTE_TIPO_FACTURA_B in allowed:
                self.cbte_combo.addItem("Factura B (RI / venta a consumidor)", CBTE_TIPO_FACTURA_B)

            if hasattr(self, "cbte_lock_label"):
                self.cbte_lock_label.setText(taxpayer_type_lock_text(taxpayer_type))
        finally:
            self.cbte_combo.blockSignals(False)

    def _refresh_after_setup(self) -> None:
        self.settings = get_all_settings(self.db_path)
        self.setWindowTitle(f"{self.settings.get('app_name','Factura')}")
        self._apply_cbte_lock()

    # ---------------- Tab Venta ----------------
    def _build_tab_venta(self):
        venta = QWidget()
        self.tabs.addTab(venta, "Venta")
        lay = QVBoxLayout(venta)

        doc_row = QHBoxLayout()
        self.rb_cf = QRadioButton("Consumidor Final (sin DNI/CUIT)")
        self.rb_doc = QRadioButton("Con DNI/CUIT")
        self.rb_cf.setChecked(True)

        self.doc_group = QButtonGroup(self)
        self.doc_group.addButton(self.rb_cf)
        self.doc_group.addButton(self.rb_doc)

        self.doc_input = QLineEdit()
        self.doc_input.setPlaceholderText("DNI (8) o CUIT (11)")
        self.doc_input.setEnabled(False)
        self.rb_doc.toggled.connect(lambda v: self.doc_input.setEnabled(bool(v)))

        # --- Tipo de comprobante (B / C) ---
        tipo_row = QHBoxLayout()
        tipo_row.addWidget(QLabel("Tipo de comprobante:"))

        self.cbte_combo = QComboBox()
        tipo_row.addWidget(self.cbte_combo, 1)

        self.cbte_lock_label = QLabel("")
        self.cbte_lock_label.setStyleSheet("color: #d8b24d;")
        self.cbte_lock_label.setWordWrap(True)

        hint = QLabel("• C: sin IVA.  • B: IVA incluido (por defecto 21%).")
        hint.setStyleSheet("color: #b0b0b0;")
        hint.setWordWrap(True)

        lay.addLayout(tipo_row)
        lay.addWidget(self.cbte_lock_label)
        lay.addWidget(hint)

        self._apply_cbte_lock()

        doc_row.addWidget(self.rb_cf)
        doc_row.addWidget(self.rb_doc)
        lay.addLayout(doc_row)
        lay.addWidget(self.doc_input)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Item", "Cant.", "Precio", "Subtotal"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.itemChanged.connect(self.recalc_all)
        lay.addWidget(self.table)

        btn_row = QHBoxLayout()
        self.btn_add = QPushButton("Agregar ítem")
        self.btn_del = QPushButton("Borrar ítem")
        self.btn_fact = QPushButton("FACTURAR + IMPRIMIR (F4)")
        btn_row.addWidget(self.btn_add)
        btn_row.addWidget(self.btn_del)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_fact)
        lay.addLayout(btn_row)

        self.total_label = QLabel("TOTAL: 0.00")
        self.total_label.setAlignment(Qt.AlignRight)
        self.total_label.setStyleSheet("font-size: 18px; font-weight: 700;")
        lay.addWidget(self.total_label)

        self.btn_add.clicked.connect(self.add_row)
        self.btn_del.clicked.connect(self.del_row)
        self.btn_fact.clicked.connect(self.on_facturar)

        self.add_row()

    def add_row(self):
        r = self.table.rowCount()
        self.table.insertRow(r)
        self.table.setItem(r, 0, QTableWidgetItem(""))
        self.table.setItem(r, 1, QTableWidgetItem("1"))
        self.table.setItem(r, 2, QTableWidgetItem("0.00"))
        sub = QTableWidgetItem("0.00")
        sub.setFlags(sub.flags() & ~Qt.ItemIsEditable)
        self.table.setItem(r, 3, sub)

    def del_row(self):
        r = self.table.currentRow()
        if r >= 0:
            self.table.removeRow(r)
            self.recalc_all()

    def _to_float(self, s: str) -> float:
        s = (s or "").strip().replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    def recalc_all(self):
        self.table.blockSignals(True)
        try:
            total = 0.0
            for r in range(self.table.rowCount()):
                qty = self._to_float(self.table.item(r, 1).text() if self.table.item(r, 1) else "0")
                price = self._to_float(self.table.item(r, 2).text() if self.table.item(r, 2) else "0")
                subtotal = max(0.0, qty) * max(0.0, price)
                total += subtotal
                it = self.table.item(r, 3)
                if it:
                    it.setText(f"{subtotal:.2f}")
            self.total_label.setText(f"TOTAL: {total:.2f}")
        finally:
            self.table.blockSignals(False)

    def gather_items(self):
        items = []
        for r in range(self.table.rowCount()):
            name = (self.table.item(r, 0).text() if self.table.item(r, 0) else "").strip()
            if not name:
                continue
            qty = self._to_float(self.table.item(r, 1).text() if self.table.item(r, 1) else "0")
            price = self._to_float(self.table.item(r, 2).text() if self.table.item(r, 2) else "0")
            subtotal = max(0.0, qty) * max(0.0, price)
            if subtotal <= 0:
                continue
            items.append({"name": name, "qty": qty, "price": price, "subtotal": subtotal})
        return items

    def on_facturar(self):
        try:
            # refrescar settings por si cambiaron en Config
            self.settings = get_all_settings(self.db_path)
            self.afip = AfipService(
                modo=self.settings.get("modo", "PROD"),
                pv=int(self.settings.get("punto_venta", "14") or 14),
                cuit=int(self.settings.get("cuit_emisor", "0") or 0),
                cert_crt_path=self.settings.get("cert_crt_path", ""),
                private_key_path=self.settings.get("private_key_path", ""),
                private_key_password=(self.settings.get("private_key_password") or None),
                openssl_path=(self.settings.get("openssl_path") or None),
            )

            items = self.gather_items()
            if not items:
                QMessageBox.warning(self, "Faltan ítems", "Cargá al menos 1 ítem con importe > 0.")
                return

            total = round(sum(i["subtotal"] for i in items), 2)

            # Doc
            if self.rb_cf.isChecked():
                doc_tipo = DOC_TIPO_SIN_DOC
                doc_nro = "0"
            else:
                doc = self.doc_input.text().strip()
                if not doc.isdigit() or len(doc) not in (8, 11):
                    QMessageBox.warning(self, "Documento inválido", "Ingresá DNI (8) o CUIT (11).")
                    return
                doc_nro = doc
                doc_tipo = DOC_TIPO_DNI if len(doc) == 8 else DOC_TIPO_CUIT
            cbte_tipo = int(self.cbte_combo.currentData())
            if not self._is_cbte_allowed(cbte_tipo):
                QMessageBox.warning(self, "Tipo de comprobante bloqueado", taxpayer_blocked_cbte_message(get_setting(self.db_path, "taxpayer_type", "MONO"), cbte_tipo))
                return
            # Emitir CAE
            #res = self.afip.emitir_factura_c(doc_tipo=doc_tipo, doc_nro=doc_nro, total=total, items=items)

            res = self.afip.emitir_comprobante(
            cbte_tipo=cbte_tipo,
            doc_tipo=doc_tipo,
            doc_nro=doc_nro,
            total=total,
            items=items
            )

            inv_id = insert_invoice(
                self.db_path,
                pv=int(self.settings.get("punto_venta", "14") or 14),
                cbte_tipo=cbte_tipo,
                cbte_nro=res.cbte_nro,
                doc_tipo=doc_tipo,
                doc_nro=doc_nro,
                imp_total=total,
                cae=res.cae,
                cae_vto=res.cae_vto,
                modo=self.settings.get("modo", "PROD"),
                items=items,
            )

            # imprimir desde data
            inv_pack = get_invoice_with_items(self.db_path, inv_id)
            print_ticket_from_data(db_path=self.db_path, inv=inv_pack["invoice"], items=inv_pack["items"])

            QMessageBox.information(self, "OK", f"Factura emitida (ID {inv_id})\nNRO: {res.cbte_nro}\nCAE: {res.cae}")
            self.table.setRowCount(0)
            self.add_row()
            self.recalc_all()
            self.refresh_invoices()
            self.refresh_range()

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ---------------- Tab Comprobantes ----------------
    def _build_tab_comprobantes(self):
        comp = QWidget()
        self.tabs.addTab(comp, "Comprobantes")
        lay = QVBoxLayout(comp)

        top = QHBoxLayout()
        lay.addLayout(top)

        self.date_pick = QDateEdit()
        self.date_pick.setCalendarPopup(True)
        self.date_pick.setDate(QDate.currentDate())
        top.addWidget(QLabel("Día:"))
        top.addWidget(self.date_pick)

        self.btn_refresh = QPushButton("Actualizar")
        top.addWidget(self.btn_refresh)
        top.addStretch(1)

        self.sum_label = QLabel("Resumen: -")
        self.sum_label.setAlignment(Qt.AlignRight)
        self.sum_label.setStyleSheet("font-size: 16px; font-weight: 700;")
        lay.addWidget(self.sum_label)

        self.inv_table = QTableWidget(0, 8)
        self.inv_table.setHorizontalHeaderLabels(["ID", "Fecha", "PV", "Nro", "Doc", "Total", "CAE", "Modo"])
        for i in range(8):
            self.inv_table.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self.inv_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.inv_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        self.inv_table.setEditTriggers(QTableWidget.NoEditTriggers)
        lay.addWidget(self.inv_table)

        self.btn_refresh.clicked.connect(self.refresh_invoices)
        self.date_pick.dateChanged.connect(lambda _: self.refresh_invoices())
        self.inv_table.cellDoubleClicked.connect(self.on_reprint_clicked)

        self.refresh_invoices()

    def refresh_invoices(self):
        d = self.date_pick.date().toString("yyyy-MM-dd")
        rows = list_invoices(self.db_path, date_yyyy_mm_dd=d, limit=500)
        summ = daily_summary(self.db_path, date_yyyy_mm_dd=d)

        self.sum_label.setText(f"Resumen {d}: {summ['cant']} comprobantes | Total: {summ['total']:.2f}")

        self.inv_table.setRowCount(0)
        for r in rows:
            rr = self.inv_table.rowCount()
            self.inv_table.insertRow(rr)

            doc = f"{r['doc_tipo']}-{r['doc_nro']}"
            self.inv_table.setItem(rr, 0, QTableWidgetItem(str(r["id"])))
            self.inv_table.setItem(rr, 1, QTableWidgetItem(str(r["created_at"])))
            self.inv_table.setItem(rr, 2, QTableWidgetItem(str(r["pv"])))
            self.inv_table.setItem(rr, 3, QTableWidgetItem(str(r["cbte_nro"])))
            self.inv_table.setItem(rr, 4, QTableWidgetItem(doc))
            self.inv_table.setItem(rr, 5, QTableWidgetItem(f"{float(r['imp_total']):.2f}"))
            self.inv_table.setItem(rr, 6, QTableWidgetItem(str(r["cae"])))
            self.inv_table.setItem(rr, 7, QTableWidgetItem(str(r["modo"])))

    def on_reprint_clicked(self, row: int, col: int):
        it = self.inv_table.item(row, 0)
        if not it:
            return
        inv_id = int(it.text())
        ok = QMessageBox.question(self, "Reimprimir", f"¿Reimprimir comprobante ID {inv_id}?", QMessageBox.Yes | QMessageBox.No)
        if ok != QMessageBox.Yes:
            return
        try:
            pack = get_invoice_with_items(self.db_path, inv_id)
            print_ticket_from_data(db_path=self.db_path, inv=pack["invoice"], items=pack["items"])
            QMessageBox.information(self, "OK", "Reimpresión enviada a la impresora.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ---------------- Tab Reporte (rango) ----------------
    def _build_tab_reporte(self):
        rep = QWidget()
        self.tabs.addTab(rep, "Reporte")
        lay = QVBoxLayout(rep)

        top = QHBoxLayout()
        lay.addLayout(top)

        self.from_pick = QDateEdit()
        self.to_pick = QDateEdit()
        for w in (self.from_pick, self.to_pick):
            w.setCalendarPopup(True)
        self.to_pick.setDate(QDate.currentDate())
        self.from_pick.setDate(QDate.currentDate().addDays(-7))

        top.addWidget(QLabel("Desde:"))
        top.addWidget(self.from_pick)
        top.addWidget(QLabel("Hasta:"))
        top.addWidget(self.to_pick)

        self.btn_range = QPushButton("Actualizar")
        self.btn_export = QPushButton("Exportar Excel")
        top.addWidget(self.btn_range)
        top.addWidget(self.btn_export)
        top.addStretch(1)

        self.range_label = QLabel("Resumen: -")
        self.range_label.setAlignment(Qt.AlignRight)
        self.range_label.setStyleSheet("font-size: 16px; font-weight: 700;")
        lay.addWidget(self.range_label)

        self.range_table = QTableWidget(0, 8)
        self.range_table.setHorizontalHeaderLabels(["ID", "Fecha", "PV", "Nro", "Doc", "Total", "CAE", "Modo"])
        for i in range(8):
            self.range_table.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self.range_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.range_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        self.range_table.setEditTriggers(QTableWidget.NoEditTriggers)
        lay.addWidget(self.range_table)

        self.btn_range.clicked.connect(self.refresh_range)
        self.from_pick.dateChanged.connect(lambda _: self.refresh_range())
        self.to_pick.dateChanged.connect(lambda _: self.refresh_range())
        self.btn_export.clicked.connect(self.export_range_excel)

        self.refresh_range()

    def refresh_range(self):
        d0 = self.from_pick.date().toString("yyyy-MM-dd")
        d1 = self.to_pick.date().toString("yyyy-MM-dd")
        rows = list_invoices(self.db_path, from_yyyy_mm_dd=d0, to_yyyy_mm_dd=d1, limit=5000)
        summ = range_summary(self.db_path, from_yyyy_mm_dd=d0, to_yyyy_mm_dd=d1)
        self.range_label.setText(f"Resumen {d0} a {d1}: {summ['cant']} comprobantes | Total: {summ['total']:.2f}")

        self.range_table.setRowCount(0)
        for r in rows:
            rr = self.range_table.rowCount()
            self.range_table.insertRow(rr)
            doc = f"{r['doc_tipo']}-{r['doc_nro']}"
            self.range_table.setItem(rr, 0, QTableWidgetItem(str(r["id"])))
            self.range_table.setItem(rr, 1, QTableWidgetItem(str(r["created_at"])))
            self.range_table.setItem(rr, 2, QTableWidgetItem(str(r["pv"])))
            self.range_table.setItem(rr, 3, QTableWidgetItem(str(r["cbte_nro"])))
            self.range_table.setItem(rr, 4, QTableWidgetItem(doc))
            self.range_table.setItem(rr, 5, QTableWidgetItem(f"{float(r['imp_total']):.2f}"))
            self.range_table.setItem(rr, 6, QTableWidgetItem(str(r["cae"])))
            self.range_table.setItem(rr, 7, QTableWidgetItem(str(r["modo"])))

    def export_range_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            d0 = self.from_pick.date().toString("yyyy-MM-dd")
            d1 = self.to_pick.date().toString("yyyy-MM-dd")
            rows = list_invoices(self.db_path, from_yyyy_mm_dd=d0, to_yyyy_mm_dd=d1, limit=100000)

            path, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", f"comprobantes_{d0}_a_{d1}.xlsx", "Excel (*.xlsx)")
            if not path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Comprobantes"

            headers = ["ID", "Fecha", "PV", "CbteTipo", "CbteNro", "DocTipo", "DocNro", "Total", "CAE", "VtoCAE", "Modo"]
            ws.append(headers)
            for r in rows:
                ws.append([
                    r["id"], r["created_at"], r["pv"], r["cbte_tipo"], r["cbte_nro"],
                    r["doc_tipo"], r["doc_nro"], float(r["imp_total"]), r["cae"], r["cae_vto"], r["modo"]
                ])

            # autosize
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

            wb.save(path)
            QMessageBox.information(self, "OK", f"Excel generado:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ---------------- Tab Config ----------------
    def _build_tab_config(self):
        cfg = QWidget()
        self.tabs.addTab(cfg, "Configuración")
        lay = QVBoxLayout(cfg)

        # Settings form (DB-backed)
        form = QFormLayout()
        lay.addLayout(form)

        self.cfg_app = QLineEdit(get_setting(self.db_path, "app_name", "LocutorioWEB"))
        self.cfg_rs = QLineEdit(get_setting(self.db_path, "razon_social", "LocutorioWEB"))
        self.cfg_cuit = QLineEdit(get_setting(self.db_path, "cuit_emisor", "0"))
        self.cfg_pv = QLineEdit(get_setting(self.db_path, "punto_venta", "14"))

        self.cfg_modo = QComboBox()
        self.cfg_modo.addItem("Producción (real)", "PROD")
        self.cfg_modo.addItem("Homologación (pruebas)", "HOMO")
        cur_modo = (get_setting(self.db_path, "modo", "PROD") or "PROD").strip().upper()
        self.cfg_modo.setCurrentIndex(0 if cur_modo == "PROD" else 1)

        self.cfg_taxpayer = QComboBox()
        self.cfg_taxpayer.addItem("Monotributo (solo Factura C)", "MONO")
        self.cfg_taxpayer.addItem("Responsable Inscripto (solo Factura B)", "RI")
        cur_taxpayer = (get_setting(self.db_path, "taxpayer_type", "MONO") or "MONO").strip().upper()
        self.cfg_taxpayer.setCurrentIndex(0 if cur_taxpayer != "RI" else 1)

        self.cfg_printer_contains = QLineEdit(get_setting(self.db_path, "printer_name_contains", ""))
        self.cfg_print_mode = QComboBox()
        self.cfg_print_mode.addItem("Térmica USB (ESC/POS)", "escpos")
        self.cfg_print_mode.addItem("Impresora Windows / PDF (GDI)", "gdi")
        cur_pm = (get_setting(self.db_path, "print_mode", "escpos") or "escpos").strip().lower()
        self.cfg_print_mode.setCurrentIndex(0 if cur_pm == "escpos" else 1)

        self.cfg_openssl = QLineEdit(get_setting(self.db_path, "openssl_path", ""))
        self.cfg_crt = QLineEdit(get_setting(self.db_path, "cert_crt_path", ""))
        self.cfg_key = QLineEdit(get_setting(self.db_path, "private_key_path", ""))
        self.cfg_keypass = QLineEdit(get_setting(self.db_path, "private_key_password", ""))
        self.cfg_keypass.setEchoMode(QLineEdit.Password)

        form.addRow("Nombre app:", self.cfg_app)
        form.addRow("Razón social:", self.cfg_rs)
        form.addRow("CUIT emisor:", self.cfg_cuit)
        form.addRow("Punto de venta:", self.cfg_pv)
        form.addRow("Modo AFIP:", self.cfg_modo)
        form.addRow("Régimen fiscal:", self.cfg_taxpayer)
        form.addRow("", QLabel("• Producción = real. Homologación = pruebas."))
        form.addRow("", QLabel("• Monotributo habilita Factura C. Responsable Inscripto habilita Factura B."))
        form.addRow("Impresora (opcional):", self.cfg_printer_contains)
        form.addRow("", QLabel("• Vacío = impresora predeterminada."))
        form.addRow("Tipo de impresión:", self.cfg_print_mode)
        form.addRow("", QLabel("• ESC/POS: térmicas. GDI: PDF / cualquier impresora Windows."))
        form.addRow("OpenSSL path:", self.cfg_openssl)
        form.addRow("Cert .crt path:", self.cfg_crt)
        form.addRow("Private key path:", self.cfg_key)
        form.addRow("Key password:", self.cfg_keypass)

        # Ticket template editor
        lay.addWidget(QLabel("Plantilla de ticket (una línea por renglón):"))
        self.ticket_edit = QTextEdit("\n".join(get_ticket_lines(self.db_path)))
        self.ticket_edit.setPlaceholderText("Una línea por renglón. Placeholders: {app_name} {pv} {cbte_nro} {fecha} {items_block} {total} {cae} ...")
        lay.addWidget(self.ticket_edit, 1)

        row = QHBoxLayout()
        lay.addLayout(row)
        self.btn_save_cfg = QPushButton("Guardar configuración")
        self.btn_test_print = QPushButton("Imprimir prueba")
        row.addStretch(1)
        row.addWidget(self.btn_test_print)
        row.addWidget(self.btn_save_cfg)

        self.btn_save_cfg.clicked.connect(self.save_config)
        self.btn_test_print.clicked.connect(self.test_print)

    def save_config(self):
        set_setting(self.db_path, "app_name", self.cfg_app.text().strip())
        set_setting(self.db_path, "razon_social", self.cfg_rs.text().strip())
        set_setting(self.db_path, "cuit_emisor", self.cfg_cuit.text().strip())
        set_setting(self.db_path, "punto_venta", self.cfg_pv.text().strip())
        set_setting(self.db_path, "modo", str(self.cfg_modo.currentData()).strip())
        set_setting(self.db_path, "taxpayer_type", str(self.cfg_taxpayer.currentData()).strip())
        set_setting(self.db_path, "printer_name_contains", self.cfg_printer_contains.text().strip())
        set_setting(self.db_path, "print_mode", str(self.cfg_print_mode.currentData()).strip())
        set_setting(self.db_path, "openssl_path", self.cfg_openssl.text().strip())
        set_setting(self.db_path, "cert_crt_path", self.cfg_crt.text().strip())
        set_setting(self.db_path, "private_key_path", self.cfg_key.text().strip())
        set_setting(self.db_path, "private_key_password", self.cfg_keypass.text())

        lines = self.ticket_edit.toPlainText().splitlines()
        save_ticket_lines(self.db_path, self.ticket_edit.toPlainText())

        QMessageBox.information(self, "OK", "Configuración guardada en la base de datos.")
        self.settings = get_all_settings(self.db_path)
        self.setWindowTitle(f"{self.settings.get('app_name','Locutorio')}")
        self._apply_cbte_lock()

    def test_print(self):
        try:
            # ticket dummy
            inv = {
                "pv": int(get_setting(self.db_path, "punto_venta", "14") or 14),
                "cbte_tipo": default_cbte_for_taxpayer(get_setting(self.db_path, "taxpayer_type", "MONO")),
                "cbte_nro": 99999999,
                "doc_tipo": DOC_TIPO_SIN_DOC,
                "doc_nro": "0",
                "imp_total": 123.45,
                "cae": "00000000000000",
                "cae_vto": datetime.now().strftime("%Y%m%d"),
                "modo": get_setting(self.db_path, "modo", "PROD"),
            }
            items = [
                {"item_name": "PRUEBA 1", "qty": 1, "price": 100.0, "subtotal": 100.0},
                {"item_name": "PRUEBA 2", "qty": 1, "price": 23.45, "subtotal": 23.45},
            ]
            print_ticket_from_data(db_path=self.db_path, inv=inv, items=items)
            QMessageBox.information(self, "OK", "Prueba enviada a la impresora.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


if __name__ == "__main__":
    app = QApplication([])
    w = MainWindow()
    w.resize(900, 650)  # vos ajustás altura a 1024x768 sin drama
    w.show()
    app.exec()
